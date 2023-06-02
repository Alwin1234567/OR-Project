 # import modules
import pandas as pd
from time import time
from Cottage import Cottage
import openpyxl
from math import exp
from random import random
pd.options.mode.chained_assignment = None

 # Planner class
class Planner():
    """
    Class that stores and maintaines all info regarding reservations and cottages
    """
    
    def __init__(self, cottages, reservations, restrictionlist = ["Class", "Face South", "Near Playground", \
                       "Close to the Centre", "Near Lake ", \
                       "Near car park", "Accessible for Wheelchair", \
                       "Child Friendly", "Dish Washer ", \
                       "Wi-Fi Coverage ", "Covered Terrace"]):
        self.start_time = time()
        self.print_time("starting Planner init")
         # Add day (0 is earliest reservation day) and departure day
        earliest_day = reservations["Arrival Date"].min()
        self.start_weekday = earliest_day.weekday()
        reservations["day"] = reservations["Arrival Date"].apply(lambda x: (x-earliest_day).days)
        reservations["final_day"] = reservations["day"].add(reservations["Length of Stay"]).add(-1)
        reservations["final_day"] = reservations['final_day'].clip(upper = reservations["day"].max())
        reservations["Length of Stay"] = reservations["final_day"].subtract(reservations["day"]).add(1)
         # Changes "# persons" for reservations to nearest higher or equal "max # pers"
        cottage_Max_pers = cottages["Max # Pers"].unique().tolist()
        cottage_Max_pers.append(0)
        cottage_Max_pers.sort()
        for i in range(len(cottage_Max_pers) - 1): reservations.loc[(cottage_Max_pers[i] < reservations["# Persons"]).multiply(reservations["# Persons"] < cottage_Max_pers[i + 1]), "# Persons"] = cottage_Max_pers[i + 1]
        
        self.df_cottages = cottages
        self.df_reservations = reservations
        self.restrictionlist = restrictionlist
        self.combinations = self.combine()
        daycount = reservations["final_day"].max() + 1
        self.cottages = dict()
        for ID in cottages["ID"].tolist(): self.cottages[ID] = Cottage(ID, daycount, earliest_day)
        self.print_time("finished Planner init")
    
    
    def print_time(self, msg = ""): print(msg + "   --- %s seconds ---" % (time() - self.start_time))
    
    
    def combine(self):
        """
        Function to retrieve a dataframe with all possible reservation-cottage combos.        

        Returns
        -------
        combined : pd.Dataframe
            Cross product of cottages and reservations where only allowed combos are left.
        """
        combined = pd.merge(self.df_reservations, self.df_cottages, how = "cross", suffixes = ("_res", "_cot"))
        combined = combined[combined["Max # Pers"] >= combined["# Persons"]]
        for restriction in self.restrictionlist:
            combined = combined[combined["{}_res".format(restriction)] <= combined["{}_cot".format(restriction)]]
        combined = combined[(combined["Cottage (Fixed)"] == 0).add(combined["Cottage (Fixed)"] == combined["ID_cot"])]
        combined["upgrade"] = (combined["Class_cot"] - combined["Class_res"] + \
                              combined["Max # Pers"] - combined["# Persons"] > 0).multiply(combined["Cottage (Fixed)"] == 0)
        
        combined["fitness"] = combined["Max # Pers"] - combined["# Persons"]
        for restriction in self.restrictionlist:
            combined["fitness"] = combined["fitness"].add(combined["{}_cot".format(restriction)] - combined["{}_res".format(restriction)])
        combined = combined.reset_index()
        combined["ID"] = combined.index
        return combined
    
    def assign_cottages(self):
        """
        Function that assigns reservations to the cottages. There is no failsafe yes.
        It assigns the reservation with the least possible cottages first and chooses 
        the cottage based on whether it is considered an upgrade and the fitness score.
        """
        self.print_time("started assigning cottages")
        order = self.combinations.groupby("ID_res")["index"].count().sort_values().index.tolist()
        for reservation_ID in order:
            cottages = self.combinations[self.combinations["ID_res"] == reservation_ID]
            cottages = cottages.sort_values(by = ["upgrade", "fitness"])[["ID_cot", "ID_res", "day", "Length of Stay", "upgrade"]]
            assigned = False
            for index, row in cottages.iterrows():
                cottage = self.cottages[row["ID_cot"]]
                if cottage.allowed_reservation((row["ID_res"], row["upgrade"]), row["day"], row["Length of Stay"]):
                    cottage.add_reservation((row["ID_res"], row["upgrade"]), row["day"], row["Length of Stay"])
                    assigned = True
                    break
            if not assigned: print("couldn't assign reservation {}".format(reservation_ID))
        self.print_time("ended assigning cottages")

    def display_cottages(self):
        """
        Function that prints all cottage planning.
        """
        self.print_time("started displaying cottages")
        for cottage in self.cottages.values(): cottage.display_days()
        self.print_time("ended displaying cottages")
    
    def read_assignements(self, assignments, remove = False):
        """
        Function that takes a series with the reservations assigned to a cottage and imports this into the class.

        Parameters
        ----------
        assignments : pd.Series
            Series with reservation_ID as index and cottage number as value.
        remove : bool, optional
            boolean that describes if the cottages should be emptied first. The default is False.
        """
        self.print_time("started reading assignments")
        if remove:
            for cottage in self.cottages.values():
                for reservation in cottage.reservations:
                    cottage.remove_reservation(reservation)
        for reservation, cottage in assignments.items():
            combination = self.combinations.loc[(self.combinations["ID_res"] == reservation).multiply(self.combinations["ID_cot"] == cottage)].squeeze()
            self.cottages[cottage].add_reservation((combination["ID_res"], combination["upgrade"]), combination["day"], combination["Length of Stay"])
        self.print_time("ended reading assignments")
    
    def reservation_assignments(self):
        """
        Function that returns a dataframe with the reservations as index and the cottage number they are assigned to.
        """
        assignments = dict()
        for cottage in self.cottages.values():
            ID = cottage.ID
            for reservation in cottage.reservations: assignments[reservation[0]] = ID
        return pd.Series(data = assignments).sort_index()
    
    def store_excel(self, filename, sheetname):
        """
        Function that stores the assigned cottage of a reservation in Excel.
        """
        repeat = True
        while repeat:
            try: 
                workbook = openpyxl.load_workbook(filename = filename)
                workbook.save(filename)
                workbook.close()
                repeat = False
            except:
                answer = input("Python can't open {}.\nIf the file is opened please close it\nDo you want to try again? (y/n)".format(filename))
                if answer != "y": return
        self.print_time("started writing in excel")
        workbook = openpyxl.load_workbook(filename = filename)
        worksheet = workbook[sheetname]
        for i, cottage in enumerate(self.reservation_assignments().values.tolist(), start = 2):
            worksheet.cell(row = i, column = 2).value = cottage
        workbook.save(filename)
        workbook.close()
        self.print_time("ended writing in excel")
    
    def assign_improvements_best(self):
        """
        Function that tries to find the best switch for each cottage and applies them. (Very slow)
        """
        self.print_time("started improving assignments with a score of {}".format(self.score))
        count = 0
        improved = True
        while improved:
            improved = False
            assignments = self.reservation_assignments()
            cottageIDs = list()
            cottagescores = list()
            for cottage in self.cottages.values():
                cottageIDs.append(cottage.ID)
                cottagescores.append(cottage.score)
            order = pd.Series(cottagescores, index = cottageIDs).sort_values(ascending = False).index.tolist()
            best = (0, 0)
            for cottage_ID in order:
                reservations = self.combinations[self.combinations["ID_cot"] == cottage_ID]
                for index, row in reservations.iterrows():
                    skip = False
                    for reservation in self.cottages[cottage_ID].reservations: 
                        if row["ID_res"] == reservation[0]: 
                            skip = True
                            break
                    if skip: continue
                    startscore = self.cottages[assignments.loc[row["ID_res"]]].score + \
                                 self.cottages[cottage_ID].score
                    if self.cottages[cottage_ID].allowed_reservation((row["ID_res"], row["upgrade"]), row["day"], row["Length of Stay"]):
                        old_cottage_ID = assignments.loc[row["ID_res"]]
                        if not self.switch_cottage(row["ID_res"], cottage_ID): print("error while switching cottages")
                        endscore = self.cottages[old_cottage_ID].score + self.cottages[cottage_ID].score
                        if not self.switch_cottage(row["ID_res"], old_cottage_ID): print("error while switching back cottages")
                        score = startscore - endscore
                        if score > best[1]:
                            best = (row["ID_res"], score)
                            improved = True
                if improved:
                    if not self.switch_cottage(best[0], cottage_ID): print("error while switching cottages")
                    break
            count += 1
            if count % 1 == 0: self.print_time("iteratie {} with score {}".format(count, self.score))
        self.print_time("ended improving assignments with a score of {}".format(self.score)) 
    
    def assign_improvements_any(self, max_time = 300):
        """
        Function that tries to find a better score and applies them (slow).

        Parameters
        ----------
        max_time : INT, optional
            Sets the maximum amount of seconds the function is allowed to run.. The default is 300.
        """
        self.print_time("started improving assignments with a score of {}".format(self.score))
        count = 0
        improved = True
        runtime = time()
        repeat_after = 5
        while time() - runtime < max_time:
            improved = False
            assignments = self.reservation_assignments()
            cottageIDs = list()
            cottagescores = list()
            for cottage in self.cottages.values():
                cottageIDs.append(cottage.ID)
                cottagescores.append(cottage.score)
            order = pd.Series(cottagescores, index = cottageIDs).sort_values(ascending = False).index.tolist()
            for i, cottage_ID in enumerate(order):
                reservations = self.combinations[self.combinations["ID_cot"] == cottage_ID]
                for index, row in reservations.iterrows():
                    skip = False
                    for reservation in self.cottages[cottage_ID].reservations: 
                        if row["ID_res"] == reservation[0]: 
                            skip = True
                            break
                    if skip: continue
                    startscore = self.cottages[assignments.loc[row["ID_res"]]].score + \
                                 self.cottages[cottage_ID].score
                    if self.cottages[cottage_ID].allowed_reservation((row["ID_res"], row["upgrade"]), row["day"], row["Length of Stay"]):
                        old_cottage_ID = assignments.loc[row["ID_res"]]
                        if not self.switch_cottage(row["ID_res"], cottage_ID): print("error while switching cottages")
                        endscore = self.cottages[old_cottage_ID].score + self.cottages[cottage_ID].score
                        score = startscore - endscore
                        if score <= 0:
                            if not self.switch_cottage(row["ID_res"], old_cottage_ID): print("error while switching back cottages")
                        else:
                            improved = True
                            count += 1
                            assignments.loc[row["ID_res"]] = cottage_ID
                            if count % 10 == 0: self.print_time("iteratiion {} with score {}".format(count, self.score))
                if i >= repeat_after:
                    if not improved: repeat_after += 1
                    else: break
        self.print_time("ended improving assignments with a score of {}".format(self.score)) 
    
    
    def assign_improvements_simulated(self, max_time = 300, temperature_init_mul = 0.0001, temperature_mul = 0.5, temperature_repeat = 10):
        """
        Function that uses simulated annealing to try and find a better score (slow).

        Parameters
        ----------
        max_time : INT, optional
            Sets the maximum amount of seconds the function is allowed to run. The default is 300.
        temperature_init_mul : FLOAT, optional
            multiplier of the first score to create initial temperature. The default is 0.0001.
        temperature_mul : FLOAT, optional
            multiplier of the temperature after each repeat. The default is 0.5.
        temperature_repeat : INT, optional
            amount of itterations before the temperature is reduced. The default is 10.
        """
        self.print_time("started improving assignments with a score of {}".format(self.score))
        temperature = self.score * temperature_init_mul
        runtime = time()
        assignments = [self.reservation_assignments(), self.score]
        best_assignment = (assignments[0].copy, self.score)
        combinations = self.combinations[self.combinations["Cottage (Fixed)"] == 0]
        itteration = 0
        while time() - runtime < max_time:
            success = False
            row = combinations.sample(replace = False).squeeze()
            if assignments[0].loc[row["ID_res"]] == row["ID_cot"] or \
                not self.cottages[row["ID_cot"]].allowed_reservation((row["ID_res"], row["upgrade"]), row["day"], row["Length of Stay"]): pass
            else:
                old_cottage_ID = assignments[0].loc[row["ID_res"]]
                startscore = self.cottages[assignments[0].loc[row["ID_res"]]].score + \
                             self.cottages[row["ID_cot"]].score
                if not self.switch_cottage(row["ID_res"], row["ID_cot"]): print("error while switching cottages")
                endscore = self.cottages[old_cottage_ID].score + self.cottages[row["ID_cot"]].score
                score = startscore - endscore
                if score < 0:
                    success = random() < exp(score / temperature)
                else: success = True
                if success:
                    assignments[0].loc[row["ID_res"]] = row["ID_cot"]
                    assignments[1] += score
                    if assignments[1] > best_assignment[1]: best_assignment = (assignments[0].copy(), self.score)
                    itteration += 1
                    if itteration % 10 == 0: self.print_time("iteration {} with score {}".format(itteration, self.score))
                    if itteration % temperature_repeat == 0: temperature *= temperature_mul
                else:
                    if not self.switch_cottage(row["ID_res"], old_cottage_ID): print("error while switching back cottages")
        self.print_time("ended improving assignments with a score of {}".format(self.score))
        self.read_assignements(best_assignment[0], remove = True)
    
    
    def gaps_legionella_optimiser_repeat(self, max_time = 600, gaps_1 = True, gaps_2 = True, gaps_3 = True, gaps_456 = True):
        """
        Function that alternates gaps and legionella improvements until time runs out or no more improvements are found.

        Parameters
        ----------
        max_time : INT, optional
            Sets the maximum amount of seconds the function is allowed to run. The default is 600.
        """
        self.print_time("started repeating gaps and legionella with a score of {}".format(self.score))
        runtime = time()
        while True:
            improved = self.gaps_optimiser(max_time = max_time + runtime - time(), gaps_1 = gaps_1, gaps_2 = gaps_2, gaps_3 = gaps_3, gaps_456 = gaps_456)
            improved += self.legionella_optimiser(max_time = max_time + runtime - time())
            if not improved:
                self.print_time("ended repeating gaps and legionella with a score of {}".format(self.score))
                return
            if time() - runtime > max_time:
                self.print_time("ended repeating gaps and legionella with a score of {}".format(self.score))
                return
    
    
    def gaps_optimiser(self, max_time = 300, gaps_1 = True, gaps_2 = True, gaps_3 = True, gaps_456 = True):
        """
        Function that goes over the gaps in cottages and tries to find a suitable reservation to fill the gap without creating new gaps.

        Parameters
        ----------
        max_time : INT, optional
            Sets the maximum amount of seconds the function is allowed to run. The default is 600.. The default is 300.

        Returns
        -------
        has_improved : BOOL
            variable that tracks if any imprvements have been made.

        """
        self.print_time("started improving gaps with a score of {}".format(self.score))
        itteration = 0
        has_improved = False
        runtime = time()
        improved = True
        while improved: # verschil tussen 1 run en meerdere is minimaal
            cottageIDs = list()
            cottagescores = list()
            improved = False
            for cottage in self.cottages.values():
                cottageIDs.append(cottage.ID)
                cottagescores.append(cottage.score)
            order = pd.Series(cottagescores, index = cottageIDs).sort_values(ascending = False).index.tolist()
            for cottage_ID in order:
                gaps = self.cottages[cottage_ID].gaps
                if gaps == 0: continue
                if gaps_3: 
                    new_itteration = self.find_gap_improvement_3_caller(gaps, cottage_ID, itteration)
                    if new_itteration > itteration:
                        itteration = new_itteration
                        has_improved = True
                compressed = self.cottages[cottage_ID].compressed_days()
                front_reservation = None
                gap = None
                last = None
                for filler in compressed:
                    if filler[0] == None:
                        improved = False
                        gap = filler
                        last = filler
                        if gaps_456:
                            if self.find_gap_improvement_456(cottage_ID, filler[1], filler[2]):
                                improved = True
                                has_improved = True
                                itteration += 1
                                if itteration % 20 == 0: self.print_time("iteration {} with score {}".format(itteration, self.score))
                        if gaps_1 and not improved and front_reservation != None:
                            if self.find_gap_improvement_1(cottage_ID, filler[1], filler[2], front_reservation):
                                improved = True
                                has_improved = True
                                itteration += 1
                                if itteration % 20 == 0: self.print_time("iteration {} with score {}".format(itteration, self.score))
                    else:
                        front_reservation = filler
                        if gap != None:
                            if gaps_2 and not improved and last[0] == None:
                                if self.find_gap_improvement_2(cottage_ID, gap[1], gap[2], filler):
                                    improved = True
                                    has_improved = True
                                    itteration += 1
                                    if itteration % 20 == 0: self.print_time("iteration {} with score {}".format(itteration, self.score))
                        last = filler
                if time() - runtime > max_time:
                    self.print_time("ended improving gaps with a score of {}".format(self.score))
                    return has_improved
        self.print_time("ended improving gaps with a score of {}".format(self.score))
        return has_improved
        
    def legionella_optimiser(self, max_time = 300):
        """
        Function that finds cottages with legionella and tries to find reservations to remove legionella without creating new gaps.

        Parameters
        ----------
        max_time : INT, optional
            Sets the maximum amount of seconds the function is allowed to run. The default is 600.. The default is 300.

        Returns
        -------
        has_improved : BOOL
            variable that tracks if any imprvements have been made..

        """
        self.print_time("started improving legionella with a score of {}".format(self.score))
        runtime = time()
        has_improved = False
        itteration = 0
        if self.legionellas == 0:
            self.print_time("ended improving legionella with a score of {}".format(self.score))
            return has_improved
        while True:
            assignments = self.reservation_assignments()
            options = self.combinations
            options["edges"] = self.combinations["ID_cot"].map(lambda row: self.cottages[row].legionella_edges)
            options = options[options["edges"].map(lambda row: row[0]) > 0]
            options["begin_edge"] = options["edges"].map(lambda row: row[1])
            options["end_edge"] = options["edges"].map(lambda row: row[2])
            options = options.drop(columns = ["edges"])
            options = options[(options["day"] < options["begin_edge"] + 22).multiply(options["day"] >= options["begin_edge"]).multiply(options["final_day"] > options["end_edge"] - 22).multiply(options["final_day"] <= options["end_edge"])]
            
            options = self.get_empty(options, options["day"] - 1, options["final_day"] + 1)
            if options.empty: 
                self.print_time("ended improving legionella with a score of {}".format(self.score))
                return has_improved
            options = options[options["ID_res"].map(lambda ID: self.cottages[assignments[ID]].remove_no_legionella(ID))]
            if options.empty: 
                self.print_time("ended improving legionella with a score of {}".format(self.score))
                return has_improved
            options = options.drop_duplicates('ID_res', keep = 'first')
            options = options.drop_duplicates('ID_cot', keep = 'first')
            for index, row in options.iterrows():
                if not self.switch_cottage(row["ID_res"], row["ID_cot"]): print("error while switching cottages")
                itteration += 1
                has_improved = True
                if itteration % 10 == 0: self.print_time("iteration {} with score {}".format(itteration, self.score))
            if time() - runtime > max_time:
                self.print_time("ended improving legionella with a score of {}".format(self.score))
                return has_improved
    
    
    def upgrade_optimiser(self, max_time = 300):
        """
        Function that finds duos of reservations that both have an upgrade. Then switches them if that reduces the amount of upgrades.

        Parameters
        ----------
        max_time : INT, optional
            Sets the maximum amount of seconds the function is allowed to run. The default is 600.. The default is 300.. The default is 300.
        """
        self.print_time("started improving upgrades with a score of {}".format(self.score))
        runtime = time()
        itteration = 0
        while True:
            assignments = self.reservation_assignments()
            assignments_merged = pd.Series(assignments.reset_index().values.tolist())
            assignments = assignments[assignments_merged.map(lambda row: self.cottages[row[1]].is_upgrade(row[0])).values]
            options = self.df_reservations
            options = options[options["ID"].isin(assignments.index)]
            options = pd.merge(options, options, how = "cross", suffixes = ("_1", "_2"))
            options = options[(options["day_1"] == options["day_2"]).multiply(options["final_day_1"] == options["final_day_2"]).multiply(options["ID_1"] != options["ID_2"])]
            options["improvement"] = options[["ID_1", "ID_2"]].values.tolist()
            options["improvement"] = options["improvement"].map(sorted)
            options["improvementstr"] = options["improvement"].map(str)
            options = options.drop_duplicates('improvementstr', keep = 'first')
            options = options[options["improvement"].map(lambda reservation_ids: self.possible_upgrade(reservation_ids[0], reservation_ids[1], assignments))]
            if options.empty:
                self.print_time("ended improving upgrades with a score of {}".format(self.score))
                return
            while not options.empty:
                reservations = options.iloc[0]
                if not self.swap_cottages(reservations["ID_1"], reservations["ID_2"]): print("error while swapping cottages")
                options = options[(options["ID_1"] != reservations["ID_1"]).multiply(options["ID_1"] != reservations["ID_2"]).multiply(options["ID_2"] != reservations["ID_1"]).multiply(options["ID_2"] != reservations["ID_2"])]
                itteration += 1
                if itteration % 10 == 0: self.print_time("iteration {} with score {}".format(itteration, self.score))
            if time() - runtime > max_time:
                self.print_time("ended improving upgrades with a score of {}".format(self.score))
                return


    def fritothugaps_optimiser(self, max_time = 300):
        self.print_time("started improving fritothugaps with a score of {}".format(self.score))
        runtime = time()
        itteration = 0
        while True:
            improved = False
            for cottage in self.cottages.values():
                if cottage.gaps == 0: continue
                compressed = cottage.compressed_days()
                front_reservation = None
                gap = None
                last = None
                for filler in compressed:
                    if filler[0] == None:
                        success = False
                        last = filler
                        if (filler[1] + self.start_weekday) % 7 in [4, 5]:
                            if front_reservation != None:
                                options = self.filter_fritothuoptions(front_reservation, filler, cottage.ID, "left")
                                print(1)
                                if not options.empty:
                                    print(2)
                                    success = self.swap_cottages(front_reservation[0][0], options.iloc[0]["ID_res"])
                                    if success: 
                                        improved = True
                                        itteration += 1
                                        if itteration % 10 == 0: self.print_time("iteration {} with score {}".format(itteration, self.score))
                                    else: print("error while swapping cottages")
                                    
                        if (not success or filler[2] - filler[1] >= 10) and (filler[2] + self.start_weekday) % 7 in [2, 3]:
                            gap = filler
                        else: gap = None
                    else:
                        if gap != None and last[0] == None:
                            options = self.filter_fritothuoptions(filler, gap, cottage.ID, "right")
                            if not options.empty:
                                print(options.iloc[0]["ID_res"], filler[0][0], cottage.ID, )
                                success = self.swap_cottages(filler[0][0], options.iloc[0]["ID_res"])
                                if success: 
                                    improved = True
                                    itteration += 1
                                    if itteration % 10 == 0: self.print_time("iteration {} with score {}".format(itteration, self.score))
                                else: print("error while swapping cottages")
                                
                        last = filler
                
                    if time() - runtime > max_time:
                        self.print_time("ended improving fritothugaps with a score of {}".format(self.score))
                        return
            if not improved:
                self.print_time("ended improving fritothugaps with a score of {}".format(self.score))
                return


    def find_gap_improvement_3_caller(self, gaps, cottage_ID, itteration):
        compressed = self.cottages[cottage_ID].compressed_days()
        for swap_set in range(gaps - 1):
            first_gap = None
            middle_reservation = None
            last = None
            for filler in compressed:
                if filler[0] == None: 
                    if first_gap == None: first_gap = filler
                    else:
                        if self.find_gap_improvement_3(cottage_ID, first_gap[1], middle_reservation[0][0], filler[2]):
                            first_gap = None
                            itteration += 1
                            if itteration % 10 == 0: self.print_time("iteration {} with score {}".format(itteration, self.score))
                        else:
                            first_gap = filler
                elif last == None:
                    middle_reservation = filler
                    last = filler
                else: 
                    first_gap = None
        return itteration


    def find_gap_improvement_1(self, cottage_ID, gap_start, gap_end, front_reservation):
        """
        Function that tries to find a suitable reservation to swap the reservation with and also fill the gap without creating a new gap.

        Parameters
        ----------
        cottage_ID : INT
            ID of cottage with gap.
        gap_start : TYPE
            first day of the gap.
        gap_end : INT
            last day of gap.
        front_reservation : INT
            ID and day of reservation in front of gap.

        Returns
        -------
        BOOL
            variable that indicates if an inporvement has been made.

        """
        combinations = self.combinations
        assignments = self.reservation_assignments()
        allowed_reservations = assignments[assignments.isin(combinations[combinations["ID_res"] == front_reservation[0][0]]["ID_cot"])].index
        options = combinations[combinations["ID_cot"] == cottage_ID]
        options = options[options["day"] == front_reservation[1]]
        options = options[options["final_day"] == gap_end]
        if options.empty: return False
        options = options[options["ID_res"].isin(allowed_reservations)]
        if options.empty: return False
        options = self.get_empty(options, front_reservation[1] - 1, gap_end + 1, side = "right")
        if options.empty: return False
        if not self.swap_cottages(front_reservation[0][0], options.iloc[0]["ID_res"]): print("error while swapping cottages")
        return True
    
    def find_gap_improvement_2(self, cottage_ID, gap_start, gap_end, back_reservation):
        """
        Function that tries to find a suitable reservation to swap the reservation with and also fill the gap without creating a new gap.

        Parameters
        ----------
        cottage_ID : INT
            ID of cottage with gap.
        gap_start : TYPE
            first day of the gap.
        gap_end : INT
            last day of gap.
        back_reservation : INT
            ID and day of reservation in after of gap.

        Returns
        -------
        BOOL
            variable that indicates if an inporvement has been made.

        """
        combinations = self.combinations
        assignments = self.reservation_assignments()
        allowed_reservations = assignments[assignments.isin(combinations[combinations["ID_res"] == back_reservation[0][0]]["ID_cot"])].index
        # options = combinations[(combinations["ID_cot"] == cottage_ID).multiply(combinations["day"] == gap_start).multiply(combinations["final_day"] == back_reservation[2]).multiply(combinations["ID_res"].isin(allowed_reservations))]
        options = combinations[combinations["ID_cot"] == cottage_ID]
        options = options[options["day"] == gap_start]
        options = options[options["final_day"] == back_reservation[2]]
        options = options[options["ID_res"].isin(allowed_reservations)]
        if options.empty: return False
        options = self.get_empty(options, gap_start - 1, back_reservation[2] + 1, side = "left")
        if options.empty: return False
        if not self.swap_cottages(back_reservation[0][0], options.iloc[0]["ID_res"]): print("error while swapping cottages")
        return True


    def find_gap_improvement_3(self, cottage_ID, first_gap_start, middle_reservation, second_gap_end):
        """
        function that swaps two gaps with a reservation in between with a reservation that covers both the gaps and the reservation.

        Parameters
        ----------
        cottage_ID : INT
            ID of cottage with gap.
        first_gap_start : INT
            first day of the first gap.
        middle_reservation : INT
            ID of reservation to be swapped.
        second_gap_end : INT
            last day of second reservation.

        Returns
        -------
        BOOL
            variable that indicates if an inporvement has been made.

        """
        assignments = self.reservation_assignments()
        combinations = self.combinations
        allowed_reservations = assignments[assignments.isin(combinations[combinations["ID_res"] == middle_reservation]["ID_cot"])].index
        # options = combinations[(combinations["ID_cot"] == cottage_ID).multiply(combinations["day"] == first_gap_start).multiply(combinations["final_day"] == second_gap_end).multiply(combinations["ID_res"].isin(allowed_reservations))]
        options = combinations[(combinations["ID_cot"] == cottage_ID)]
        options = options[options["day"] == first_gap_start]
        options = options[options["final_day"] == second_gap_end]
        if options.empty: return False
        options = options[options["ID_res"].isin(allowed_reservations)]
        if options.empty: return False
        options = self.get_empty(options, first_gap_start - 1, second_gap_end + 1, side = "both")
        if options.empty: return False
        if not self.swap_cottages(options.iloc[0]["ID_res"], middle_reservation):
            print("error while swapping cottages")
            return False
        return True
        
        
    def find_gap_improvement_456(self, cottage_ID, gap_start, gap_end):
        """
        Function that tries to find a reservation to fill the gap in the cottage without creating a new gap.

        Parameters
        ----------
        cottage_ID : INT
            ID of cottage with gap.
        gap_start : INT
            first day of the gap.
        gap_end : INT
            last day of gap.

        Returns
        -------
        BOOL
            variable that indicates if an inporvement has been made.

        """
        combinations = self.combinations
        options = combinations[(combinations["ID_cot"] == cottage_ID).multiply(combinations["day"] == gap_start).multiply(combinations["final_day"] == gap_end)]
        if options.empty: return False
        options = self.get_empty(options, gap_start - 1, gap_end + 1, side = "both")
        if options.empty: return False
        if not self.switch_cottage(options.iloc[0]["ID_res"], cottage_ID):
            print("error while switching cottages")
            return False
        return True
        

    def get_empty(self, options, start, end, side = "both", reverse = False):
        """
        Function that filters for reservations with gaps next to them.

        Parameters
        ----------
        options : pd.DataFrame
            DataFrame with reservations.
        start : INT or pd.Series
            Value that indicates the day on which it needs to check for a gap. Usually on the left side of a reservation.
        end : INT or pd.Series
            Value that indicates the day on which it needs to check for a gap. Usually on the right side of a reservation.
        side : str ("left" or "right" or "both"), optional
            Indicates which side it needs to find a gap. The default is "both".

        Returns
        -------
        options : pd.DataFrame
            filtered DataFrame where only rows with reservations that have at least one gap next to them are left.

        """
        if side not in ["left", "right", "both"]:
            print("get_empty: invalid side {}".format(side))
            return options
        assignments = self.reservation_assignments()
        if side == "left" or side == "both":
            options["left_empty"] = start
            options["left_empty"] = options[["ID_res", "left_empty"]].values.tolist()
            if reverse: options["left_empty"] = ~options["left_empty"].map(lambda row: self.cottages[assignments[row[0]]].empty_day(row[1]))
            else: options["left_empty"] = options["left_empty"].map(lambda row: self.cottages[assignments[row[0]]].empty_day(row[1]))
        if side == "right" or side == "both":
            options["right_empty"] = end
            options["right_empty"] = options[["ID_res", "right_empty"]].values.tolist()
            if reverse: options["right_empty"] = ~options["right_empty"].map(lambda row: self.cottages[assignments[row[0]]].empty_day(row[1]))
            else: options["right_empty"] = options["right_empty"].map(lambda row: self.cottages[assignments[row[0]]].empty_day(row[1]))
        
        if side == "left":
            options = options[options["left_empty"]]
            return options
        if side == "right":
            options = options[options["right_empty"]]
            return options
        if side == "both":
            options = options[options["left_empty"].add(options["right_empty"])]
            if options.empty: return options
            options = options.sort_values(by = ["left_empty", "right_empty"], ascending = False)
            return options


    def possible_upgrade(self, ID_res1, ID_res2, assignments):
        """
        Function that tests if a swap between the reservations is possible and will reduce the amount of upgrades.

        Parameters
        ----------
        ID_res1 : INT
            ID of first reservation.
        ID_res2 : INT
            ID of second reservation.
        assignments : Series
            self.reservation_assignments.

        Returns
        -------
        BOOL
            variable that indicates if switching the reservations will be benificial.

        """
        combinations = self.combinations
        cottage1 = assignments.loc[ID_res1]
        cottage2 = assignments.loc[ID_res2]
        # combi1 = combinations.loc[(combinations["ID_res"] == ID_res1).multiply(combinations["ID_cot"] == cottage2)]
        # combi2 = combinations.loc[(combinations["ID_res"] == ID_res2).multiply(combinations["ID_cot"] == cottage1)]
        combi1 = combinations[(combinations["ID_res"] == ID_res1).multiply(combinations["ID_cot"] == cottage2)]
        combi2 = combinations[(combinations["ID_res"] == ID_res2).multiply(combinations["ID_cot"] == cottage1)]
        if combi1.empty or combi2.empty: return False
        # return not bool(combi1.squeeze()["upgrade"] * combi2.squeeze()["upgrade"])
        return not bool(combi1["upgrade"].sum() * combi2["upgrade"].sum())


    def filter_fritothuoptions(self, reservation, gap, cottage_ID, side):
        assignments = self.reservation_assignments()
        combinations = self.combinations
        allowed_reservations = assignments[assignments.isin(combinations[combinations["ID_res"] == reservation[0][0]]["ID_cot"])].index
        options = combinations[combinations["ID_cot"] == cottage_ID]
        
        if side == "right":
            options = options[options["day"] == reservation[1]]
            options = options[options["final_day"] < gap[1] - (gap[1] + self.start_weekday) % 7 - 3]
            options = options[options["final_day"] > gap[2] - 21]
            options = self.get_empty(options, reservation[1], options["final_day"] + 1, side = "right")
            if options.empty: return options
            options = self.get_empty(options, reservation[1], options["final_day"] + 7, side = "right", reverse = True)
        else:
            options = options[options["final_day"] == reservation[2]]
            options = options[options["day"] > gap[2] + (reservation[2] + self.start_weekday) % 7 - 1]
            options = options[options["day"] < gap[1] + 21]
            options = self.get_empty(options, options["day"] - 1, reservation[2], side = "left")
            if options.empty: return options
            options = self.get_empty(options, options["day"] - 7, reservation[2], side = "left", reverse = True)
        if options.empty: return options
        options = options[options["ID_res"].isin(allowed_reservations)]
        return options       


    def switch_cottage(self, ID_res, new_cottage_ID):
        """
        Function that puts the reservation in the cottage if possible.

        Parameters
        ----------
        ID_res : INT
            ID of the reservation that will be switched.
        new_cottage_ID : INT
            ID of the cottage the reservation is going to.

        Returns
        -------
        BOOL
            Variable that indicates if the switch is succesfull.

        """
        combinations = self.combinations
        assignments = self.reservation_assignments()
        reservation_old = self.cottages[assignments.loc[ID_res]].find_reservation(ID_res)
        reservation_new = combinations.loc[(combinations["ID_res"] == ID_res).multiply(combinations["ID_cot"] == new_cottage_ID)].squeeze()
        if self.cottages[new_cottage_ID].allowed_reservation((reservation_new["ID_res"], reservation_new["upgrade"]), reservation_new["day"], reservation_new["Length of Stay"]):
            self.cottages[assignments.loc[ID_res]].remove_reservation(reservation_old)
            self.cottages[new_cottage_ID].add_reservation((reservation_new["ID_res"], reservation_new["upgrade"]), reservation_new["day"], reservation_new["Length of Stay"])
            return True
        return False
    
    def swap_cottages(self, ID_res1, ID_res2):
        """
        Function that swaps two reservations if possible

        Parameters
        ----------
        ID_res1 : INT
            ID of fthe first reservation.
        ID_res2 : INT
            ID of fthe second reservation..

        Returns
        -------
        BOOL
            Indicates id the swap was succesfull.

        """
        combinations = self.combinations
        assignments = self.reservation_assignments()
        cottage1 = assignments.loc[ID_res1]
        cottage2 = assignments.loc[ID_res2]
        reservation_old1 = self.cottages[cottage1].find_reservation(ID_res1)
        reservation_old2 = self.cottages[cottage2].find_reservation(ID_res2)
        reservation_new1 = combinations.loc[(combinations["ID_res"] == ID_res1).multiply(combinations["ID_cot"] == cottage2)].squeeze()
        reservation_new2 = combinations.loc[(combinations["ID_res"] == ID_res2).multiply(combinations["ID_cot"] == cottage1)].squeeze()
        if type(reservation_new1) != pd.Series or type(reservation_new2) != pd.Series: return False
        self.cottages[cottage1].remove_reservation(reservation_old1)
        self.cottages[cottage2].remove_reservation(reservation_old2)
        if self.cottages[cottage2].allowed_reservation((reservation_new1["ID_res"], reservation_new1["upgrade"]), reservation_new1["day"], reservation_new1["Length of Stay"]) and \
           self.cottages[cottage1].allowed_reservation((reservation_new2["ID_res"], reservation_new2["upgrade"]), reservation_new2["day"], reservation_new2["Length of Stay"]):
               self.cottages[cottage2].add_reservation((reservation_new1["ID_res"], reservation_new1["upgrade"]), reservation_new1["day"], reservation_new1["Length of Stay"])
               self.cottages[cottage1].add_reservation((reservation_new2["ID_res"], reservation_new2["upgrade"]), reservation_new2["day"], reservation_new2["Length of Stay"])
               return True
        self.cottages[cottage1].add_reservation((reservation_new1["ID_res"], reservation_new1["upgrade"]), reservation_new1["day"], reservation_new1["Length of Stay"])
        self.cottages[cottage2].add_reservation((reservation_new2["ID_res"], reservation_new2["upgrade"]), reservation_new2["day"], reservation_new2["Length of Stay"])
        return False
        
    def results(self):
        """
        Function that prints the total score and the individual scores of the current assignment.
        """        
        msg = "The current assignments have a score of {}\n".format(self.score)
        msg += "The scores are as follows:\n"
        msg += "{} gaps for a score of {}\n".format(self.gaps, self.gaps * 6)
        msg += "{} legionella gaps for a score of {}\n".format(self.legionellas, self.legionellas * 12)
        msg += "{} fri to thu gaps for a score of {}\n".format(self.fritothus, self.fritothus * -3)
        msg += "{} upgardes for a score of {}".format(self.upgrades, self.upgrades * 1)
        print(msg)
        
        


    @property
    def score(self):
        total = 0
        for cottage in self.cottages.values(): total += cottage.score
        return total
    
    @property
    def upgrades(self):
        total = 0
        for cottage in self.cottages.values(): total += cottage.upgrades
        return total
    
    @property
    def gaps(self):
        total = 0
        for cottage in self.cottages.values(): total += cottage.gaps
        return total
    
    @property
    def legionellas(self):
        total = 0
        for cottage in self.cottages.values(): total += cottage.legionellas
        return total
    
    @property
    def fritothus(self):
        total = 0
        for cottage in self.cottages.values(): total += cottage.fritothus
        return total